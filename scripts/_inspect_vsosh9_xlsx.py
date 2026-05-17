# -*- coding: utf-8 -*-
"""Quick inspection of vsosh9_methods.xlsx structure."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('vsosh9_methods.xlsx', read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name!r}  (dims: {ws.max_row} rows x {ws.max_column} cols) ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        cells = [(str(c)[:60] if c is not None else '') for c in row[:14]]
        print(f"  row{i}: " + " | ".join(cells))
    if ws.max_column > 14:
        print(f"  (+ {ws.max_column - 14} more cols truncated)")
