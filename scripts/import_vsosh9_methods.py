# -*- coding: utf-8 -*-
"""Импорт банка ВсОШ-9 методов и задач из vsosh9_methods.xlsx.

Идемпотентный импортёр: при повторном запуске обновляет существующие записи,
ничего не дублирует.

Что делает:
  1. Лист «89 методов»     → upsert TheoryBlock (code, name, section, total_count).
  2. Лист «Топ методов»    → обновляет total_count + share_percent + frequency_vsosh_9.
  3. Лист «Все 295 задач»  → создаёт по одному архивному Probnik на (год+этап) ВсОШ-9
                            и заливает все 295 задач как OlympiadTask с привязкой
                            к методу через method_primary/method_secondary.

Лист «ВсОШ 9 — методы»    → не нужен на стороне БД, поскольку та же связь уже
                            хранится в OlympiadTask.method_primary/secondary.
Лист «Сводка год × этап»  → справочный, не импортируется (агрегаты дёшево считать
                            on-the-fly в шаблоне).

Usage:
    python scripts/import_vsosh9_methods.py            # импорт из vsosh9_methods.xlsx
    python scripts/import_vsosh9_methods.py --dry-run  # сухой прогон (rollback в конце)
    python scripts/import_vsosh9_methods.py --path X   # явный путь к xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Iterable, Optional

import openpyxl

# Гарантируем, что корень проекта в sys.path для `import app, models`.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from app import app  # noqa: E402
from models import db  # noqa: E402
from models_olympiad import (  # noqa: E402
    Probnik,
    OlympiadTask,
    TheoryBlock,
    THEORY_SECTIONS,
)


# ── Mapping: первая буква кода → раздел (категория) метода ───────────────────
# A=алгебра, B=логика, C=алгебра, D=теория чисел, E=комбинаторика,
# F=геометрия, G=алгебра, H=прочее. Соответствует ТЗ.
SECTION_BY_LETTER = {
    'A': 'A', 'B': 'B', 'C': 'C', 'D': 'D',
    'E': 'E', 'F': 'F', 'G': 'G', 'H': 'H',
}

# Этап → префикс задачи (один символ для компактного OlympiadTask.number).
STAGE_PREFIX = {
    'Школьный':      'Ш',
    'Муниципальный': 'М',
    'Региональный':  'Р',
    'Заключительный':'З',
}

# Этап → нормализованный slug для OlympiadTask.stage (английская техническая метка).
STAGE_NORMALIZED = {
    'Школьный':       'school',
    'Муниципальный':  'municipal',
    'Региональный':   'regional',
    'Заключительный': 'final',
}


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _safe_int(v) -> Optional[int]:
    if v is None or v == '':
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _section_for_code(code: str) -> Optional[str]:
    if not code:
        return None
    letter = code[0].upper()
    sec = SECTION_BY_LETTER.get(letter)
    # TheoryBlock.section валидируется enum THEORY_SECTIONS — гарантируем совместимость.
    if sec and sec in THEORY_SECTIONS:
        return sec
    return None


# ── stage 1: «89 методов» ────────────────────────────────────────────────────

def import_methods_sheet(ws) -> dict:
    """Upsert TheoryBlock по коду метода. Возвращает {code: TheoryBlock}.

    Колонки листа: Код | Название | Раздел | В ВсОШ 9
    """
    stats = {'created': 0, 'updated': 0, 'skipped': 0}
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], rows[1:]
    print(f"  Sheet '89 методов': {len(body)} rows")

    by_code: dict[str, TheoryBlock] = {}
    for row in body:
        code = _safe_str(row[0]) if len(row) >= 1 else ''
        name = _safe_str(row[1]) if len(row) >= 2 else ''
        section_human = _safe_str(row[2]) if len(row) >= 3 else ''
        count = _safe_int(row[3]) if len(row) >= 4 else None
        if not code or not name:
            stats['skipped'] += 1
            continue

        section = _section_for_code(code)
        tb = TheoryBlock.query.filter_by(method_code=code).first()
        if tb is None:
            tb = TheoryBlock(
                method_code=code,
                method_name=name,
                section=section,
                total_count=count,
                frequency_vsosh_9=count,  # как best-effort альтернатива
            )
            db.session.add(tb)
            stats['created'] += 1
        else:
            # idempotent update
            tb.method_name = name
            if section and not tb.section:
                tb.section = section
            if count is not None:
                tb.total_count = count
                if tb.frequency_vsosh_9 is None:
                    tb.frequency_vsosh_9 = count
            stats['updated'] += 1

        by_code[code] = tb

    db.session.flush()  # получаем id у только что созданных
    print(f"  TheoryBlock: created={stats['created']}, updated={stats['updated']}, "
          f"skipped={stats['skipped']}")
    return by_code


# ── stage 2: «Топ методов» — total_count + share_percent ─────────────────────

def import_top_methods_sheet(ws, by_code: dict) -> int:
    """Обновляет TheoryBlock.{total_count, share_percent, frequency_vsosh_9}
    из таблицы «Топ методов».

    Колонки: # | Код | Название | Кол-во задач | Доля
    """
    updated = 0
    rows = list(ws.iter_rows(values_only=True))
    body = rows[1:]
    print(f"  Sheet 'Топ методов': {len(body)} rows")

    for row in body:
        if len(row) < 5:
            continue
        code = _safe_str(row[1])
        count = _safe_int(row[3])
        share = row[4]
        if not code:
            continue

        tb = by_code.get(code) or TheoryBlock.query.filter_by(method_code=code).first()
        if tb is None:
            continue  # код, которого нет в «89 методов» — пропускаем

        if count is not None:
            tb.total_count = count
            tb.frequency_vsosh_9 = count
        if share is not None:
            try:
                tb.share_percent = float(share)
            except (TypeError, ValueError):
                pass
        updated += 1

    print(f"  TOP stats applied to {updated} methods")
    return updated


# ── stage 3: «Все 295 задач» → Probnik + OlympiadTask ────────────────────────

def _ensure_archive_probnik(year: int, sort_idx: int) -> Probnik:
    """Возвращает (или создаёт) архивный Probnik для конкретного года ВсОШ-9.

    Все этапы (Муниципальный/Региональный/Заключительный) одного года кладутся
    в один Probnik, тип='topic', number = year_offset для уникальности.
    """
    season_year = year  # храним фактический год, чтобы не путать с 2027.
    type_ = 'topic'
    # number в Probnik должен быть уникален среди (competition, grade, season_year, type).
    # У нас на год = 1 архивный пробник — number=1.
    number = 1
    code = f'vsosh-9-archive-{year}'

    p = Probnik.query.filter_by(code=code).first()
    if p is None:
        p = Probnik(
            code=code,
            type=type_,
            number=number,
            title=f'ВсОШ-9 архив, {year}',
            description=f'Реальные задачи заключительного, регионального и муниципального этапов ВсОШ 9 класса за {year} год.',
            competition='ВсОШ',
            grade=9,
            season_year=season_year,
            sort_order=sort_idx,
            is_published=False,  # архив, не основной контент
        )
        db.session.add(p)
        db.session.flush()
    return p


def import_tasks_sheet(ws) -> dict:
    """Заливает 295 задач в OlympiadTask, идемпотентно по (probnik, number).

    Колонки: # | Год | Этап | День | № | Метод | Доп. | Уверенность | Текст (фрагмент)
    """
    stats = {'created': 0, 'updated': 0, 'skipped': 0, 'probniks': 0}
    rows = list(ws.iter_rows(values_only=True))
    body = rows[1:]
    print(f"  Sheet 'Все 295 задач': {len(body)} rows")

    probnik_by_year: dict[int, Probnik] = {}
    year_sort_idx = 0
    sort_within_probnik: dict[int, int] = {}

    for row in body:
        if len(row) < 9:
            stats['skipped'] += 1
            continue
        year = _safe_int(row[1])
        stage = _safe_str(row[2])
        day = _safe_str(row[3])  # 1-й / 2-й день заключительного этапа
        num_in_stage = _safe_str(row[4])
        method_primary = _safe_str(row[5])
        method_secondary = _safe_str(row[6]) or None
        text_excerpt = _safe_str(row[8])

        if not year or not stage or not num_in_stage or not method_primary:
            stats['skipped'] += 1
            continue

        prefix = STAGE_PREFIX.get(stage, '?')
        # Учитываем «День» для заключительного этапа (иначе теряем 30 задач
        # из-за коллизии номеров 1..5 на двух днях).
        if day:
            digits = ''.join(ch for ch in day if ch.isdigit())
            if digits:
                prefix = f'{prefix}{digits[0]}'
        # Уникальный номер задачи внутри Probnik: "М-1", "Р-3", "З1-5", "З2-5".
        task_number = f'{prefix}-{num_in_stage}'

        if year not in probnik_by_year:
            year_sort_idx += 1
            probnik_by_year[year] = _ensure_archive_probnik(year, year_sort_idx)
            stats['probniks'] += 1

        probnik = probnik_by_year[year]
        sort_within_probnik.setdefault(probnik.id, 0)
        sort_within_probnik[probnik.id] += 1
        sort_order = sort_within_probnik[probnik.id]

        existing = OlympiadTask.query.filter_by(
            probnik_id=probnik.id, number=task_number
        ).first()

        # Условие задачи — у нас только фрагмент текста; для архивного материала ок.
        condition_md = text_excerpt or '*Текст задачи будет добавлен позже.*'
        idea_md = '*Идея решения будет добавлена позже.*'
        solution_md = '*Полное решение будет добавлено позже.*'

        # method_codes — нормализованный JSON-список всех методов (без дублей,
        # без пустых строк). Используется для фильтра «задачи, в которых
        # встречается метод X» через JSON contains / Python.
        codes_list = [method_primary]
        if method_secondary and method_secondary != method_primary:
            codes_list.append(method_secondary)
        # Дополнительные методы могут разделяться запятыми / пробелами / `;`.
        if method_secondary:
            for piece in (
                method_secondary.replace(';', ',').replace('/', ',').split(',')
            ):
                piece = piece.strip()
                if piece and piece not in codes_list:
                    codes_list.append(piece)

        stage_slug = STAGE_NORMALIZED.get(stage)

        if existing is None:
            task = OlympiadTask(
                probnik_id=probnik.id,
                number=task_number,
                sort_order=sort_order,
                method_primary=method_primary,
                method_secondary=method_secondary,
                method_codes=codes_list,
                year=year,
                stage=stage_slug,
                condition_md=condition_md,
                idea_md=idea_md,
                solution_md=solution_md,
                source_prototype=f'ВсОШ-9 / {stage} / {year} / №{num_in_stage}',
            )
            db.session.add(task)
            stats['created'] += 1
        else:
            existing.method_primary = method_primary
            existing.method_secondary = method_secondary
            existing.method_codes = codes_list
            existing.year = year
            existing.stage = stage_slug
            existing.sort_order = sort_order
            existing.source_prototype = (
                f'ВсОШ-9 / {stage} / {year} / №{num_in_stage}'
            )
            # Не перетираем содержательные поля, если они уже наполнены вручную.
            if existing.condition_md.startswith('*') or not existing.condition_md.strip():
                existing.condition_md = condition_md
            stats['updated'] += 1

    print(
        f"  Archive Probniks created: {stats['probniks']} (по одному на год)"
    )
    print(
        f"  OlympiadTask: created={stats['created']}, "
        f"updated={stats['updated']}, skipped={stats['skipped']}"
    )
    return stats


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Импорт банка ВсОШ-9 методов')
    parser.add_argument(
        '--path', default='vsosh9_methods.xlsx',
        help='Путь к xlsx (по умолчанию vsosh9_methods.xlsx в корне проекта)',
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Откатить транзакцию в конце (только для проверки)',
    )
    args = parser.parse_args()

    if not os.path.isfile(args.path):
        print(f'❌ Файл не найден: {args.path}', file=sys.stderr)
        sys.exit(1)

    print(f'📂 Открываю {args.path}')
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    required = {'89 методов', 'Топ методов', 'Все 295 задач'}
    missing = required - set(wb.sheetnames)
    if missing:
        print(f'❌ В xlsx нет листов: {missing}', file=sys.stderr)
        sys.exit(2)

    with app.app_context():
        print('🧮 Stage 1: «89 методов» → TheoryBlock')
        by_code = import_methods_sheet(wb['89 методов'])

        print('📈 Stage 2: «Топ методов» → total_count + share_percent')
        import_top_methods_sheet(wb['Топ методов'], by_code)

        print('📝 Stage 3: «Все 295 задач» → Probnik + OlympiadTask')
        import_tasks_sheet(wb['Все 295 задач'])

        if args.dry_run:
            db.session.rollback()
            print('🟡 Dry-run: транзакция откачена.')
        else:
            db.session.commit()
            print('✅ Импорт завершён, изменения сохранены.')


if __name__ == '__main__':
    main()
